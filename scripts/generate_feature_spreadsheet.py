"""
Generate SAT Feature Planning & Review Spreadsheet.
Run: python3 scripts/generate_feature_spreadsheet.py
Output: docs/SAT_Feature_Planning.xlsx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent.parent / "docs" / "SAT_Feature_Planning.xlsx"
OUT.parent.mkdir(exist_ok=True)

# ── Palette ────────────────────────────────────────────────────────────────────
C = {
    "header_bg":    "1E3A5F",
    "header_fg":    "FFFFFF",
    "subheader_bg": "2E6DA4",
    "subheader_fg": "FFFFFF",
    "done":         "D4EDDA",
    "done_border":  "28A745",
    "in_progress":  "FFF3CD",
    "ip_border":    "FFC107",
    "backlog":      "F8D7DA",
    "bl_border":    "DC3545",
    "na":           "F0F0F0",
    "na_border":    "AAAAAA",
    "alt_row":      "F5F8FF",
    "white":        "FFFFFF",
    "section_bg":   "D6E4F0",
    "critical":     "FF4444",
    "high":         "FF8800",
    "medium":       "FFCC00",
    "low":          "88CC44",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color="888888"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, values, bg=C["header_bg"], fg=C["header_fg"], height=22):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=10)
        c.alignment = align("center", "center", wrap=True)
        c.border = border("AAAAAA")

def data_row(ws, row, values, bg=C["white"], height=18, wrap_cols=None):
    ws.row_dimensions[row].height = height
    wrap_cols = wrap_cols or []
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(size=9)
        c.alignment = align("left", "center", wrap=(col in wrap_cols))
        c.border = border()

def status_fill(status):
    s = (status or "").strip().lower()
    if s in ("done", "complete", "completed", "yes"):
        return fill(C["done"])
    if s in ("in progress", "in_progress", "partial"):
        return fill(C["in_progress"])
    if s in ("backlog", "not started", "no", "tbd", "none", "—"):
        return fill(C["backlog"])
    return fill(C["na"])

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def add_filter(ws, ref):
    ws.auto_filter.ref = ref

def section_title(ws, row, title, ncols, bg=C["section_bg"]):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = fill(bg)
    c.font = font(bold=True, size=11, color="1E3A5F")
    c.alignment = align("left", "center")
    c.border = thick_border()
    ws.row_dimensions[row].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════

FEATURES = [
    # epic, us_id, story_short, jira, jira_status, code_status, owner, pts, priority,
    # source_repo, commit, service_dir, contract, flag, fvd, notes
    (
        "E1", "US-01", "Wind Analysis & Ventilation",
        "SAT-4", "Done", "POC — not merged to org",
        "Vishwas", 13, "Must Have",
        "Vishwas721/sat (sat04/)", "55cd75e", "services/wind/",
        "contracts/wind.yaml", "feature.wind.climatology",
        "FVD-09", "POC in personal repo. Must merge to org before integration."
    ),
    (
        "E1", "US-01b", "Cyclone / Special Wind Events",
        "—", "No ticket", "No code",
        "Vishwas", 8, "Must Have",
        "—", "—", "services/wind/ (extend)",
        "TBD", "feature.wind.cyclone",
        "—", "No Jira ticket. No code. Needs IMD cyclone track or IBTrACS dataset."
    ),
    (
        "E1", "US-02", "Temperature & Thermal Profile",
        "SAT-9", "Done", "Done — local backend",
        "Vishwas", 8, "Must Have",
        "Site-Analysis-Tool/src/Backend/Temperature/", "local",
        "services/temperature/",
        "contracts/temperature.yaml", "feature.temperature.thermal-profile",
        "FVD-08", "First candidate for integration. Backend fully tested."
    ),
    (
        "E1", "US-03", "Rainfall & Water Management",
        "—", "No ticket", "Unverified — Karthik claims Done",
        "Karthik", 5, "Must Have",
        "Unknown", "—", "services/ (new: rainfall)",
        "TBD", "feature.rainfall.analysis",
        "—", "Backend claimed Done in spreadsheet but code not located. Must find before integration."
    ),
    (
        "E1", "US-04", "Solar / Sun Path",
        "SAT-21", "Done (spike)", "Done — local backend",
        "Karthik", 13, "Must Have",
        "Sprint0_User_Stories + Site-Analysis-Tool/src/Backend/SunPath/",
        "944c5b8", "services/sunpath/",
        "contracts/sunpath.yaml", "feature.sunpath.diagram",
        "FVD-06", "Backend done. Shadow casting and solstice export not implemented."
    ),
    (
        "E1", "US-05", "Contour Maps & Slope Orientation",
        "—", "No ticket", "No code",
        "—", 8, "Must Have",
        "—", "—", "services/geo/ (extend)",
        "TBD", "feature.geo.slope",
        "—", "MERIT DEM already used in flood service — slope calc is a natural extension."
    ),
    (
        "E1", "US-06", "Visual & Physical Site Characteristics",
        "—", "No ticket", "No code",
        "—", 13, "Must Have",
        "—", "—", "TBD",
        "TBD", "feature.geo.visual",
        "—", "Needs design. Possible: Street View API, Mapbox GL, or satellite imagery."
    ),
    (
        "E1", "US-07", "Flood Risk / Floodplain Mapping",
        "SAT-18 (related)", "Done (spike)", "Done — local backend",
        "Tanmay", 5, "Should Have",
        "SiteAnalysis_GEE + Site-Analysis-Tool/src/Backend/FloodPlains/",
        "edd4c29", "services/flood/",
        "contracts/flood.yaml", "feature.flood.risk-analysis",
        "FVD-04", "4-component scoring done. Historical flood event data not integrated."
    ),
    (
        "E1", "—", "GEE Interactive Map (platform)",
        "SAT-18 (related)", "Done (spike)", "Done — GEE org repo",
        "—", "—", "Must Have",
        "Site-Analysis/SiteAnalysis_GEE", "0815bc7",
        "services/geo/", "TBD", "feature.geo.map",
        "FVD-01", "Core GEE platform. Feeds into geo service."
    ),
    (
        "E1", "—", "Vegetation Analysis (NDVI/EVI/SAVI)",
        "SAT-18", "Done (spike)", "Done — GEE org repo",
        "—", "—", "Must Have",
        "Site-Analysis/SiteAnalysis_GEE", "13afab7",
        "services/geo/", "TBD", "feature.geo.vegetation",
        "FVD-02", "NDVI + EVI + SAVI + MODIS time-series. Part of geo service."
    ),
    (
        "E1", "—", "Administrative Boundaries",
        "—", "No ticket", "Done — GEE org repo",
        "—", "—", "Must Have",
        "Site-Analysis/SiteAnalysis_GEE", "13afab7",
        "services/geo/", "TBD", "feature.geo.admin",
        "FVD-03", "FAO GAUL 3-tier hierarchy. Part of geo service."
    ),
    (
        "E1", "—", "PDF Report Generation",
        "SAT-23 (related)", "Done (spike)", "Done — GEE org repo",
        "—", "—", "Should Have",
        "Site-Analysis/SiteAnalysis_GEE", "edd4c29",
        "services/ (report)", "TBD", "feature.report.pdf",
        "FVD-05", "ReportLab PDF. Needs to pull data from all other services."
    ),
    (
        "E1", "—", "Next.js Frontend Platform",
        "SAT-23 (related)", "Done (spike)", "Done — SiteAnalysisV2 org repo",
        "—", "—", "Must Have",
        "Site-Analysis/SiteAnalysisV2", "0ba0e17",
        "apps/web/", "N/A (frontend)", "—",
        "FVD-07", "Next.js 15 + Firebase + GEE/OSM/KGIS/Bhuvan. apps/web/ bootstrapped."
    ),
    (
        "E2", "US-08", "Zoning Classification",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.regulatory.zoning",
        "—", "No planning done. Possible source: Bhuvan LULC / local municipal APIs."
    ),
    (
        "E2", "—", "Land Use Compliance",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.regulatory.land-use",
        "—", "Dependent on US-08 zoning data."
    ),
    (
        "E2", "US-09", "Accessibility Standards",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.regulatory.accessibility",
        "—", "UDCPR / NBC compliance rules. Likely rule-based engine, no external data source."
    ),
    (
        "E2", "US-10", "Environmental Clearance (ECBC/IGBC/GRIHA)",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.regulatory.env-clearance",
        "—", "Green building compliance checklist. Needs regulatory data source."
    ),
    (
        "E2", "US-11", "CRZ / Eco-Zone Verification",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.regulatory.crz",
        "—", "SAT-22 spike done but no implementation. CRZ boundary data from MoEFCC."
    ),
    (
        "E2", "US-12", "Viewpoints & Landmarks",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.geo.viewpoints",
        "—", "3D visual cones needed. OSM landmark data available."
    ),
    (
        "E3", "US-13", "Historical Land-Use",
        "—", "No ticket", "No code",
        "—", None, "Should Have",
        "—", "—", "TBD", "TBD", "feature.history.land-use",
        "—", "Historical satellite or cadastral data. Google Earth Engine time-series possible."
    ),
    (
        "E3", "US-14", "Contamination & Heritage",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.history.heritage",
        "—", "Heritage listings from ASI / state heritage boards. Contamination: no clear data source."
    ),
    (
        "E3", "US-15", "Population Density",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.socio.population",
        "—", "WorldPop or Census India 2011 raster data. GEE can serve this."
    ),
    (
        "E3", "US-16", "Cultural & Religious Hubs",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.socio.cultural",
        "—", "OSM amenity tags (place_of_worship, community_centre, park). Available now."
    ),
    (
        "E3", "US-17", "Schools & Hospitals Proximity",
        "—", "No ticket", "No code",
        "—", None, "Must Have",
        "—", "—", "TBD", "TBD", "feature.socio.services",
        "—", "OSM amenity=school/hospital + isochrone. OSMnx already used in Django backends."
    ),
    (
        "E3", "US-18", "Local Materials",
        "—", "No ticket", "No code",
        "—", None, "Should Have",
        "—", "—", "TBD", "TBD", "feature.socio.materials",
        "—", "No external data source identified. Likely curated static dataset by region."
    ),
]

INTEGRATION_ORDER = [
    # order, feature, service, contract, flag, from_repo, to_path, prereqs, owner, sprint, status, blocker
    (1, "Temperature & Thermal Profile", "services/temperature/", "contracts/temperature.yaml",
     "feature.temperature.thermal-profile",
     "Site-Analysis-Tool/src/Backend/Temperature/",
     "services/temperature/app/",
     "None", "Vishwas", "Sprint 3", "Ready", "None"),
    (2, "Solar / Sun Path", "services/sunpath/", "contracts/sunpath.yaml",
     "feature.sunpath.diagram",
     "Site-Analysis-Tool/src/Backend/SunPath/",
     "services/sunpath/app/",
     "None", "Karthik", "Sprint 3", "Ready", "None"),
    (3, "Flood Risk", "services/flood/", "contracts/flood.yaml",
     "feature.flood.risk-analysis",
     "Site-Analysis-Tool/src/Backend/FloodPlains/",
     "services/flood/app/",
     "None", "Tanmay", "Sprint 3", "Ready", "None"),
    (4, "Wind Analysis", "services/wind/", "contracts/wind.yaml",
     "feature.wind.climatology",
     "Vishwas721/sat (sat04/)",
     "services/wind/app/",
     "None", "Vishwas", "Sprint 3", "Blocked", "POC in personal repo — must push to org first"),
    (5, "Geo / GEE Platform", "services/geo/", "contracts/geo.yaml (TBD)",
     "feature.geo.map",
     "Site-Analysis/SiteAnalysis_GEE",
     "services/geo/app/",
     "Temperature, Flood done", "TBD", "Sprint 4", "Needs contract", "Write contracts/geo.yaml first"),
    (6, "Vegetation (NDVI/EVI/SAVI)", "services/geo/ (extend)", "contracts/geo.yaml",
     "feature.geo.vegetation",
     "Site-Analysis/SiteAnalysis_GEE",
     "services/geo/app/gee_utils.py",
     "Geo platform done", "TBD", "Sprint 4", "Blocked by #5", "Depends on geo service"),
    (7, "Admin Boundaries", "services/geo/ (extend)", "contracts/geo.yaml",
     "feature.geo.admin",
     "Site-Analysis/SiteAnalysis_GEE",
     "services/geo/app/gee_utils.py",
     "Geo platform done", "TBD", "Sprint 4", "Blocked by #5", "Depends on geo service"),
    (8, "Rainfall Analysis", "services/ (new)", "contracts/rainfall.yaml (TBD)",
     "feature.rainfall.analysis",
     "Unknown — locate Karthik's code",
     "services/rainfall/",
     "Temperature done", "Karthik", "Sprint 4", "Blocked", "Code not found — locate before planning"),
    (9, "Contour / Slope", "services/geo/ (extend)", "contracts/geo.yaml",
     "feature.geo.slope",
     "New — extend MERIT DEM from flood service",
     "services/geo/app/",
     "Geo + Flood done", "TBD", "Sprint 5", "Not started", "Extend existing MERIT DEM integration"),
    (10, "PDF Reports", "services/ (report)", "contracts/report.yaml (TBD)",
     "feature.report.pdf",
     "Site-Analysis/SiteAnalysis_GEE",
     "services/report/",
     "All data services done", "TBD", "Sprint 5", "Not started", "Depends on all upstream services"),
    (11, "Frontend — Map & Analysis UI", "apps/web/", "N/A",
     "All feature flags",
     "Site-Analysis/SiteAnalysisV2",
     "apps/web/",
     "At least 1 service done", "Vishwas / Dhanush", "Sprint 3+", "In progress", "apps/web bootstrapped. Wire service clients."),
    (12, "Zoning / Regulatory (E2)", "services/ (new)", "TBD",
     "feature.regulatory.*",
     "None", "services/regulatory/", "None", "TBD", "Sprint 6+", "Not started", "No planning. Create Jira epic first."),
    (13, "Socio-Cultural (E3)", "services/ (new)", "TBD",
     "feature.socio.*",
     "None", "services/socio/", "None", "TBD", "Sprint 7+", "Not started", "No planning. Create Jira epic first."),
]

OPEN_ISSUES = [
    # feature, type, description, severity, owner, status
    ("Wind Analysis (US-01)",
     "Code Location",
     "POC at Vishwas721/sat (personal repo). Not in Site-Analysis org. Cannot be PR'd until pushed to org.",
     "Critical", "Vishwas", "Open"),
    ("Rainfall Analysis (US-03)",
     "Code Missing",
     "Karthik's backend marked Done in product spreadsheet but code not found in any repo. Must locate before Sprint 4.",
     "Critical", "Karthik", "Open"),
    ("Cyclone Events (US-01b)",
     "No Jira Ticket",
     "Feature in product backlog (8 pts, Must Have) with no Jira ticket, no code, no data source identified.",
     "High", "Vishwas", "Open"),
    ("All E2 Stories (US-08 to US-12)",
     "No Planning",
     "5 Must Have stories with no Jira tickets, no owners, no code, no data sources identified. Target release June 1.",
     "Critical", "Chirag", "Open"),
    ("All E3 Stories (US-13 to US-18)",
     "No Planning",
     "6 stories (4 Must Have) with no Jira tickets, no owners, no code. Target release June 1.",
     "High", "Chirag", "Open"),
    ("Sprint Planning",
     "Process Gap",
     "Sprint sheet ends at Sprint 2 (Dec 2025). No Sprint 3+ planning documented. 5 months of work untracked.",
     "High", "Chirag", "Open"),
    ("Spreadsheet Status Mismatch",
     "Data Inconsistency",
     "US-01 (Wind) and US-02 (Temperature) show 'In Progress' in product spreadsheet but are Done in Jira (SAT-4, SAT-9).",
     "Medium", "Chirag", "Open"),
    ("CODEOWNERS",
     "Repo Config",
     "@Site-Analysis/maintainers team in CODEOWNERS not configured. PRs have no enforced reviewers.",
     "Medium", "Chirag", "Open"),
    ("Branch Protection",
     "Repo Config",
     "No branch protection rules on main. Anyone can push directly, bypassing CI and PR template.",
     "High", "Chirag", "Open"),
    ("Sun Path — Shadow Casting",
     "Incomplete Feature",
     "US-04 AC: 'Seasonal shadow patterns visualized' — not implemented in current backend.",
     "Medium", "Karthik", "Open"),
    ("Sun Path — Solstice Export",
     "Incomplete Feature",
     "US-04 AC: 'Solstice data exported for design reference' — JSON available but no dedicated export format.",
     "Low", "Karthik", "Open"),
    ("Flood Risk — Historical Data",
     "Incomplete Feature",
     "US-07 AC: 'Historical flood data integrated' — current model uses hydrology proxy, not actual event data.",
     "Medium", "Tanmay", "Open"),
    ("Geo Service Contract",
     "Missing Contract",
     "contracts/geo.yaml not written. Geo service (GEE, OSM, Bhuvan, vegetation, admin) has no OpenAPI spec.",
     "High", "TBD", "Open"),
    ("Rainfall Contract",
     "Missing Contract",
     "contracts/rainfall.yaml not written. Cannot integrate rainfall feature without contract.",
     "High", "Karthik", "Open"),
    ("Wind Frontend",
     "Missing Component",
     "No windApi.ts in apps/web/. Equivalent of temperatureApi.ts not created.",
     "Medium", "Vishwas", "Open"),
    ("SAT_Backend Repo",
     "Unclear Ownership",
     "Site-Analysis/SAT_Backend repo exists (private, empty). Unclear if this is separate from services/ in SAT monorepo.",
     "Medium", "Chirag", "Open"),
    ("SAT_Feature-Agents Repo",
     "Unclear Ownership",
     "Site-Analysis/SAT_Feature-Agents repo exists (private, empty). AI agent work scope not defined.",
     "Medium", "Sreedhana", "Open"),
    ("Product Vision",
     "Documentation",
     "Product Overview sheet has vision = 0.01 (placeholder). No defined product vision statement.",
     "Low", "Chirag", "Open"),
]

BUILD_EVIDENCE = [
    # feature, commit, repo, date, jira, key_files, key_functions
    ("GEE Interactive Map", "0815bc7", "Site-Analysis/SiteAnalysis_GEE", "2025-10-05",
     "SAT-18",
     "app/main.py, app/gee_utils.py",
     "initialize_gee(), POST /analyze-location"),
    ("Vegetation Analysis", "13afab7", "Site-Analysis/SiteAnalysis_GEE", "2025-10-09",
     "SAT-18",
     "app/gee_utils.py",
     "calculate_ndvi(), calculate_evi(), calculate_savi(), get_modis_vegetation_timeseries()"),
    ("Administrative Boundaries", "13afab7", "Site-Analysis/SiteAnalysis_GEE", "2025-10-09",
     "—",
     "app/gee_utils.py",
     "get_administrative_boundaries() — FAO GAUL L1/L2/L3"),
    ("Next.js Frontend Platform", "0ba0e17", "Site-Analysis/SiteAnalysisV2", "2025-10-14",
     "SAT-23",
     "app/main.py, app/schemas/requests.py, app/schemas/responses.py, app/services/bhuvan_service.py",
     "POST /analyze — asyncio.gather([gee, osm, bhuvan, kgis])"),
    ("Solar / Sun Path", "944c5b8", "Site-Analysis/Sprint0_User_Stories", "2025-11-09",
     "SAT-21",
     "server.py",
     "create_polar_sun_path(), create_cartesian_sun_path(), calculate_solar_events(), get_timezone()"),
    ("Flood Risk Analysis", "edd4c29", "Site-Analysis/SiteAnalysis_GEE", "2025-11-25",
     "—",
     "app/gee_utils.py, app/pdf_generator.py, app/chart_generator.py",
     "calculate_flood_risk() — 4 components: MERIT DEM + hydro + JRC + MODIS LULC"),
    ("PDF Report Generation", "edd4c29", "Site-Analysis/SiteAnalysis_GEE", "2025-11-25",
     "SAT-23",
     "app/pdf_generator.py, app/chart_generator.py",
     "generate_pdf_report(), POST /generate-report"),
    ("Temperature & Thermal Profile", "local only", "Site-Analysis-Tool (local)", "2026-03-26",
     "SAT-9",
     "app/routers/weather.py, app/services/climate_analytics.py, app/services/imd_weather_service.py, app/models/climate.py",
     "GET /weather/thermal-profile, ClimateAnalyticsService.get_annual_thermal_profile()"),
    ("Wind Analysis", "55cd75e", "Vishwas721/sat (sat04/)", "2026-03-19",
     "SAT-4",
     "sat_wind_poc.py, backend_poc.py, benchmark_performance.py",
     "GET /analysis/wind/climatology, fetch_gee_wind_data(), build_rose(), get_imd_season(), get_orientation_advice()"),
]


def make_feature_registry(wb):
    ws = wb.create_sheet("1. Feature Registry")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1E3A5F"

    headers = [
        "Epic", "Story ID", "Feature / Story", "Jira", "Jira Status",
        "Code Status", "Owner", "Points", "Priority",
        "Source Repo", "Commit", "SAT Service Dir",
        "Contract", "Feature Flag", "FVD", "Notes / Blockers"
    ]
    widths = [6, 8, 28, 10, 14, 22, 10, 7, 10, 32, 10, 20, 22, 30, 8, 40]
    set_col_widths(ws, widths)
    header_row(ws, 1, headers)
    freeze(ws, "A2")
    add_filter(ws, f"A1:{get_column_letter(len(headers))}1")

    current_epic = None
    r = 2
    for row_data in FEATURES:
        epic = row_data[0]
        jira_status = row_data[4]
        code_status = row_data[5]
        bg = C["alt_row"] if r % 2 == 0 else C["white"]
        data_row(ws, r, list(row_data), bg=bg, height=20, wrap_cols=[3, 10, 16])

        # Color the Jira status cell
        jira_cell = ws.cell(row=r, column=5)
        jira_cell.fill = status_fill(jira_status)

        # Color the code status cell
        code_cell = ws.cell(row=r, column=6)
        cs = code_status.lower()
        if "done" in cs and "not" not in cs:
            code_cell.fill = fill(C["done"])
        elif "poc" in cs or "partial" in cs or "unverified" in cs or "merged" in cs:
            code_cell.fill = fill(C["in_progress"])
        else:
            code_cell.fill = fill(C["backlog"])

        # Bold epic column when it changes
        if epic != current_epic:
            ws.cell(row=r, column=1).font = font(bold=True, size=9)
            current_epic = epic
        r += 1

    ws.cell(row=1, column=1).comment = None
    return ws


def make_integration_plan(wb):
    ws = wb.create_sheet("2. Integration Plan")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2E6DA4"

    headers = [
        "#", "Feature", "Target Service",
        "Contract", "Feature Flag",
        "Source Repo / Path", "Target Path",
        "Prerequisites", "Owner", "Sprint",
        "Status", "Blocker / Next Action"
    ]
    widths = [4, 26, 20, 24, 30, 34, 22, 22, 12, 10, 14, 40]
    set_col_widths(ws, widths)
    header_row(ws, 1, headers)
    freeze(ws, "A2")
    add_filter(ws, f"A1:{get_column_letter(len(headers))}1")

    status_colors = {
        "ready":            C["done"],
        "in progress":      C["in_progress"],
        "blocked":          C["backlog"],
        "not started":      "F0F0F0",
        "needs contract":   "FFE5B4",
        "blocked by #5":    C["backlog"],
    }

    for i, row_data in enumerate(INTEGRATION_ORDER, 2):
        bg = C["alt_row"] if i % 2 == 0 else C["white"]
        data_row(ws, i, list(row_data), bg=bg, height=20, wrap_cols=[2, 6, 12])
        status = (row_data[10] or "").lower()
        for sc, scol in status_colors.items():
            if sc in status:
                ws.cell(row=i, column=11).fill = fill(scol)
                break
        # Bold the number
        ws.cell(row=i, column=1).font = font(bold=True, size=9)
    return ws


def make_open_issues(wb):
    ws = wb.create_sheet("3. Open Issues")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "DC3545"

    headers = [
        "Feature", "Issue Type", "Description",
        "Severity", "Owner", "Status", "Resolution / Notes"
    ]
    widths = [24, 18, 50, 10, 10, 10, 34]
    set_col_widths(ws, widths)
    header_row(ws, 1, headers)
    freeze(ws, "A2")
    add_filter(ws, f"A1:{get_column_letter(len(headers))}1")

    sev_colors = {
        "critical": "FFCCCC",
        "high":     "FFE0B2",
        "medium":   "FFF9C4",
        "low":      "E8F5E9",
    }

    for i, row_data in enumerate(OPEN_ISSUES, 2):
        bg = C["alt_row"] if i % 2 == 0 else C["white"]
        values = list(row_data) + [""]  # blank resolution col
        data_row(ws, i, values, bg=bg, height=20, wrap_cols=[3, 7])
        sev = (row_data[3] or "").lower()
        ws.cell(row=i, column=4).fill = fill(sev_colors.get(sev, C["white"]))
        ws.cell(row=i, column=4).font = font(bold=True, size=9)
        ws.cell(row=i, column=6).fill = status_fill(row_data[5])
    return ws


def make_build_evidence(wb):
    ws = wb.create_sheet("4. Build Evidence")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "28A745"

    headers = [
        "Feature", "Commit", "Repository", "Date Merged",
        "Jira Ticket", "Key Files", "Key Functions / Endpoints"
    ]
    widths = [24, 10, 30, 12, 12, 40, 50]
    set_col_widths(ws, widths)
    header_row(ws, 1, headers)
    freeze(ws, "A2")

    for i, row_data in enumerate(BUILD_EVIDENCE, 2):
        bg = C["alt_row"] if i % 2 == 0 else C["white"]
        data_row(ws, i, list(row_data), bg=bg, height=24, wrap_cols=[6, 7])
    return ws


def make_discussion_log(wb):
    ws = wb.create_sheet("5. Discussion Log")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FFC107"

    headers = [
        "Date", "Feature / Area", "Topic / Question",
        "Decision / Outcome", "Owner", "Due Date", "Status"
    ]
    widths = [12, 20, 40, 40, 12, 12, 12]
    set_col_widths(ws, widths)
    header_row(ws, 1, headers)
    freeze(ws, "A2")
    add_filter(ws, f"A1:{get_column_letter(len(headers))}1")

    # Pre-fill with known open questions
    open_qs = [
        ("2026-05-25", "Wind Analysis (US-01)", "Vishwas: when will sat04/ POC be pushed to Site-Analysis org?", "", "Vishwas", "", "Open"),
        ("2026-05-25", "Rainfall (US-03)", "Karthik: where is the rainfall backend code? Claimed Done in spreadsheet.", "", "Karthik", "", "Open"),
        ("2026-05-25", "Cyclone (US-01b)", "Does this need a separate Jira ticket and sprint? Who owns data sourcing (IMD/IBTrACS)?", "", "Vishwas", "", "Open"),
        ("2026-05-25", "E2 Regulatory", "Should E2 stories be deferred past June 1 release? None have code, owners, or data sources.", "", "Chirag", "2026-05-30", "Open"),
        ("2026-05-25", "SAT_Backend repo", "Is SAT_Backend a separate deployment or should everything live in SAT/services/?", "", "Chirag", "", "Open"),
        ("2026-05-25", "SAT_Feature-Agents", "What is the AI agents scope? Sreedhana: define what goes in this repo vs SAT/services/.", "", "Sreedhana", "", "Open"),
        ("2026-05-25", "Branch Protection", "Enable branch protection on main + require 1 reviewer before merging?", "", "Chirag", "2026-05-27", "Open"),
        ("2026-05-25", "Release Scope", "What is the June 1 release scope? E1 only? Which E1 stories are in/out?", "", "Chirag", "2026-05-27", "Open"),
        ("2026-05-25", "Geo Service", "Who owns geo service integration? SiteAnalysis_GEE code needs porting.", "", "TBD", "", "Open"),
        ("2026-05-25", "PDF Reports", "PDF report should aggregate all services. Define report schema before integration.", "", "TBD", "", "Open"),
    ]
    for i, row_data in enumerate(open_qs, 2):
        bg = C["alt_row"] if i % 2 == 0 else C["white"]
        data_row(ws, i, list(row_data), bg=bg, height=22, wrap_cols=[3, 4])
        ws.cell(row=i, column=7).fill = status_fill(row_data[6])

    # Add empty rows for team to fill
    for i in range(len(open_qs) + 2, len(open_qs) + 22):
        bg = C["alt_row"] if i % 2 == 0 else C["white"]
        data_row(ws, i, [""] * 7, bg=bg, height=18)
    return ws


def make_sprint_plan(wb):
    ws = wb.create_sheet("6. Sprint Plan")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "6F42C1"

    sprints = [
        ("Sprint 0", "Spike", "2025-10-27", "2025-11-10", None,
         "SAT-18, SAT-19, SAT-20, SAT-21, SAT-22, SAT-23",
         "Done", "All spikes resolved. GEE, Solar, Vegetation, Flood, PDF, Admin Boundaries explored."),
        ("Sprint 1", "Sprint", "2025-11-14", "2025-11-28", 60,
         "US-01 to US-07",
         "Partial", "US-04 (solar), US-07 (flood) resolved. US-01, US-02 slipped."),
        ("Sprint 2", "Sprint", "2025-12-03", "2025-12-17", 47,
         "SAT-04, SAT-08, SAT-11, SAT-66, US-06",
         "Partial", "SAT-4 (wind) + SAT-9 (temperature) resolved March 2026 — 3 months late."),
        ("Sprint 3 (planned)", "Sprint", "2026-06-02", "2026-06-16", 40,
         "Integrate: Temperature, SunPath, Flood into SAT monorepo. Wind after repo push.",
         "Not started", "First integration sprint. All 3 backends ready. Set up CI smoke tests."),
        ("Sprint 4 (planned)", "Sprint", "2026-06-17", "2026-06-30", 40,
         "Integrate: Wind, Geo platform, Vegetation, Admin Boundaries. Find Rainfall code.",
         "Not started", "Geo contract must be written first. Rainfall blocker must be resolved."),
        ("Sprint 5 (planned)", "Sprint", "2026-07-01", "2026-07-14", 35,
         "PDF Reports, Contour/Slope, Cyclone (if ticket created), Frontend wiring.",
         "Not started", "Aggregation sprint. PDF depends on all data services being up."),
        ("Sprint 6+ (TBD)", "Sprint", "TBD", "TBD", None,
         "E2 Regulatory stories (US-08 to US-12) — if scope confirmed.",
         "Not started", "No planning done. Create Jira epic and assign owners first."),
        ("Sprint 7+ (TBD)", "Sprint", "TBD", "TBD", None,
         "E3 Socio-Cultural stories (US-13 to US-18).",
         "Not started", "No planning done. Create Jira epic and assign owners first."),
    ]

    headers = [
        "Sprint", "Type", "Start", "End", "Velocity (pts)",
        "Stories / Scope", "Status", "Notes"
    ]
    widths = [16, 10, 12, 12, 14, 50, 12, 40]
    set_col_widths(ws, widths)
    header_row(ws, 1, headers)
    freeze(ws, "A2")

    for i, row_data in enumerate(sprints, 2):
        bg = C["alt_row"] if i % 2 == 0 else C["white"]
        data_row(ws, i, list(row_data), bg=bg, height=24, wrap_cols=[6, 8])
        ws.cell(row=i, column=7).fill = status_fill(row_data[6])
        ws.cell(row=i, column=1).font = font(bold=True, size=9)
    return ws


def make_dashboard(wb):
    ws = wb.create_sheet("0. Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF6600"
    set_col_widths(ws, [2, 22, 18, 22, 18, 22, 18, 2])
    ws.row_dimensions[1].height = 14

    # Title
    ws.merge_cells("B2:G3")
    c = ws.cell(row=2, column=2, value="SAT — Feature Planning & Review")
    c.fill = fill(C["header_bg"])
    c.font = font(bold=True, color="FFFFFF", size=16)
    c.alignment = align("center", "center")
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 16

    ws.merge_cells("B4:G4")
    c = ws.cell(row=4, column=2, value="Site Analysis Tool — Integration Tracker   |   Generated: 2026-05-25   |   Target Release: 2026-06-01")
    c.fill = fill("2E6DA4")
    c.font = font(color="FFFFFF", size=10, italic=True)
    c.alignment = align("center", "center")
    ws.row_dimensions[4].height = 18

    # Stats
    stats = [
        ("Total User Stories", "20", C["header_bg"]),
        ("Stories Done (Jira)", "4", C["done_border"]),
        ("Stories In Progress", "4", C["ip_border"]),
        ("Stories Not Started", "12", C["bl_border"]),
        ("Features with Code", "9", "28A745"),
        ("Open Issues", str(len(OPEN_ISSUES)), "DC3545"),
        ("Contracts Written", "4 / 6 needed", "2E6DA4"),
        ("Sprint 3 Ready", "3 services", "28A745"),
    ]
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 40
    ws.row_dimensions[8].height = 20

    ws.merge_cells("B6:G6")
    c = ws.cell(row=6, column=2, value="STATUS SNAPSHOT")
    c.fill = fill(C["section_bg"])
    c.font = font(bold=True, size=11, color="1E3A5F")
    c.alignment = align("center", "center")

    col_pairs = [(2, 3), (4, 5), (6, 7)]
    for idx, (label, val, color) in enumerate(stats):
        r = 7 + (idx // 3) * 3
        c1, c2 = col_pairs[idx % 3]
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=f"{label}\n{val}")
        cell.fill = fill(color if len(color) == 6 else color)
        cell.font = font(bold=True, color="FFFFFF", size=12)
        cell.alignment = align("center", "center", wrap=True)
        cell.border = thick_border()
        ws.row_dimensions[r].height = 44

    # Legend
    r_leg = 16
    ws.merge_cells(f"B{r_leg}:G{r_leg}")
    c = ws.cell(row=r_leg, column=2, value="LEGEND")
    c.fill = fill(C["section_bg"])
    c.font = font(bold=True, size=10, color="1E3A5F")
    c.alignment = align("center", "center")
    ws.row_dimensions[r_leg].height = 18

    legend = [
        (C["done"], "Done — Jira resolved + code found"),
        (C["in_progress"], "In Progress — partial / POC / unverified"),
        (C["backlog"], "Not Started — no code, no ticket"),
    ]
    for i, (color, label) in enumerate(legend):
        r = r_leg + 1 + i
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2).fill = fill(color)
        ws.cell(row=r, column=2).border = border()
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"D{r}:G{r}")
        lc = ws.cell(row=r, column=4, value=label)
        lc.font = font(size=10)
        lc.alignment = align("left", "center")

    # Navigation
    r_nav = r_leg + len(legend) + 2
    ws.merge_cells(f"B{r_nav}:G{r_nav}")
    c = ws.cell(row=r_nav, column=2, value="SHEETS")
    c.fill = fill(C["section_bg"])
    c.font = font(bold=True, size=10, color="1E3A5F")
    c.alignment = align("center", "center")
    ws.row_dimensions[r_nav].height = 18

    nav = [
        ("1. Feature Registry", "All 20 stories — status, owner, code location, contract, flag, FVD"),
        ("2. Integration Plan", "Ordered integration queue — 13 items, prerequisites, sprint, blockers"),
        ("3. Open Issues", f"{len(OPEN_ISSUES)} open issues — severity, owner, status"),
        ("4. Build Evidence", "9 features — commit hashes, repos, key files, functions"),
        ("5. Discussion Log", "Team Q&A — decisions, owners, due dates"),
        ("6. Sprint Plan", "Sprint 0–7 — scope, velocity, status, notes"),
    ]
    for i, (sheet, desc) in enumerate(nav):
        r = r_nav + 1 + i
        ws.merge_cells(f"B{r}:C{r}")
        nc = ws.cell(row=r, column=2, value=sheet)
        nc.font = font(bold=True, size=9, color="1E3A5F")
        nc.alignment = align("left", "center")
        nc.border = border()
        ws.merge_cells(f"D{r}:G{r}")
        dc = ws.cell(row=r, column=4, value=desc)
        dc.font = font(size=9)
        dc.alignment = align("left", "center")
        dc.border = border()
        ws.row_dimensions[r].height = 18

    return ws


# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

make_dashboard(wb)
make_feature_registry(wb)
make_integration_plan(wb)
make_open_issues(wb)
make_build_evidence(wb)
make_discussion_log(wb)
make_sprint_plan(wb)

wb.save(OUT)
print(f"Saved: {OUT}")
